# apps/inscricoes/management/commands/exportar_inscritos_lattes.py
# Banco de Talentos — Polo de Inovação IFG
# Exporta um .xlsx das inscrições enviadas (completa, em análise ou aprovada)
# em 3 abas — Servidores, Colaboradores Externos e Estudantes — com dados de
# perfil, URL do Lattes e os subtotais da API IFGProduz (titulação, produção,
# orientações, bancas) já consultados para quem tem Lattes. Ordenação nas
# abas: Pesquisadores primeiro, depois Apoio Técnico, alfabético em cada
# grupo. Planilha de trabalho do RH para indicação de especialistas EMBRAPII.
#
# SEM CPF de propósito (minimização LGPD): é documento interno de trabalho,
# não publicação de convocados — nome + e-mail identificam o suficiente.
#
# Uso:
#   python manage.py exportar_inscritos_lattes
#   python manage.py exportar_inscritos_lattes --saida x.xlsx
#
# Consulta a API IFGProduz (~1s por candidato com Lattes) e o banco.
# NUNCA altera dados do banco.

import time

from openpyxl import Workbook
from openpyxl.styles import Font

from django.core.management.base import BaseCommand
from django.utils import timezone

from apps.inscricoes.models import Inscricao, StatusInscricao
from apps.inscricoes.services import buscar_dados_ifgproduz
from apps.usuarios.models import Vinculo

CABECALHO = [
    "Nome",
    "Categoria",
    "Titulação",
    "Área de atuação",
    "E-mail",
    "URL Lattes",
    "Subtotal Titulação (A)",
    "Subtotal Produção (B)",
    "Subtotal Orientações (C)",
    "Subtotal Bancas (D)",
    "Total API",
    "Total autodeclarado",
    "Total validado",
    "Status inscrição",
    "Status API",
]

LARGURAS = [40, 18, 28, 30, 32, 42, 20, 20, 22, 20, 14, 18, 16, 16, 12]

ABAS = [
    (Vinculo.SERVIDOR, "Servidores"),
    (Vinculo.COLABORADOR_EXTERNO, "Colaboradores Externos"),
    (Vinculo.ESTUDANTE, "Estudantes"),
]

# Pesquisadores antes de Apoio Técnico; sem categoria por último.
ORDEM_TIPO = {"pesquisador": 0, "apoio_tecnico": 1}


class Command(BaseCommand):
    help = (
        "Gera um .xlsx em 3 abas (Servidores, Colaboradores Externos, "
        "Estudantes) das inscrições enviadas, com subtotais da API IFGProduz "
        "para quem tem Lattes, para apoiar a indicação de especialistas."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--saida",
            default=f"inscritos_lattes_{timezone.localdate()}.xlsx",
            help=(
                "Caminho do arquivo .xlsx a gerar "
                "(padrão: inscritos_lattes_<AAAA-MM-DD>.xlsx no diretório atual)."
            ),
        )

    def handle(self, *args, **options):
        # PENDENTE fica de fora (rascunho não enviado); INDEFERIDA idem.
        inscritos = list(
            Inscricao.objects.filter(
                status__in=[
                    StatusInscricao.COMPLETA,
                    StatusInscricao.EM_ANALISE,
                    StatusInscricao.APROVADA,
                ]
            ).select_related("usuario")
        )

        self.stdout.write("")
        self.stdout.write(
            self.style.MIGRATE_HEADING(f"Inscrições enviadas: {len(inscritos)}")
        )
        if not inscritos:
            self.stdout.write("Nenhuma inscrição enviada — nada a exportar.")
            return

        wb = Workbook()
        wb.remove(wb.active)
        resumo = {}  # modalidade -> {"ok": n, "falha": n, "sem_lattes": n}

        for vinculo, titulo_aba in ABAS:
            grupo = sorted(
                (i for i in inscritos if i.modalidade == vinculo),
                key=lambda i: (
                    ORDEM_TIPO.get(i.tipo_servidor, 2),
                    i.usuario.nome_completo.casefold(),
                ),
            )
            ws = wb.create_sheet(titulo_aba)
            ws.append(CABECALHO)
            for celula in ws[1]:
                celula.font = Font(bold=True)
            for coluna, largura in zip("ABCDEFGHIJKLMNO", LARGURAS):
                ws.column_dimensions[coluna].width = largura

            contagem = {"ok": 0, "falha": 0, "sem_lattes": 0}
            resumo[titulo_aba] = contagem

            for ins in grupo:
                usuario = ins.usuario
                dados, status_api = None, "sem Lattes"
                if usuario.lattes:
                    dados = buscar_dados_ifgproduz(usuario.lattes)
                    status_api = "ok" if dados else "falha"
                    time.sleep(1)  # não martelar a API institucional
                contagem["ok" if status_api == "ok" else
                         "falha" if status_api == "falha" else "sem_lattes"] += 1
                self.stdout.write(
                    f"  [{titulo_aba}] {usuario.nome_completo} — {status_api}"
                )

                def sub(chave):
                    if not dados:
                        return None  # célula vazia = preencher na mão
                    try:
                        return float(dados[chave])
                    except (KeyError, TypeError, ValueError):
                        return None

                ws.append(
                    [
                        usuario.nome_completo,
                        ins.get_tipo_servidor_display() or "—",
                        usuario.nivel_formacao or "—",
                        usuario.area_atuacao or "—",
                        usuario.email,
                        usuario.lattes or "SEM LATTES",
                        sub("subtotalA"),
                        sub("subtotalB"),
                        sub("subtotalC"),
                        sub("subtotalD"),
                        sub("total"),
                        float(ins.total) if ins.total is not None else None,
                        float(ins.total_validado) if ins.total_validado is not None else None,
                        ins.get_status_display(),
                        status_api,
                    ]
                )

        caminho = options["saida"]
        wb.save(caminho)

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("Resumo da consulta à API:"))
        for aba, c in resumo.items():
            self.stdout.write(
                f"  {aba}: {c['ok']} ok, {c['falha']} falha(s), "
                f"{c['sem_lattes']} sem Lattes"
            )
        self.stdout.write("")
        self.stdout.write(
            self.style.SUCCESS(
                f"Planilha gerada: {caminho} ({len(inscritos)} inscrição(ões))."
            )
        )
