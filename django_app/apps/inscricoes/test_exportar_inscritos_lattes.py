# apps/inscricoes/test_exportar_inscritos_lattes.py
# Banco de Talentos — Polo de Inovação IFG
# Testes do command exportar_inscritos_lattes: 3 abas, ordenação
# (Pesquisador → Apoio Técnico, alfabético) e subtotais da API (mockada).

import tempfile
from pathlib import Path
from unittest.mock import patch

from django.core.management import call_command
from django.test import TestCase
from openpyxl import load_workbook

from apps.inscricoes.models import Inscricao, StatusInscricao, TipoServidor
from apps.usuarios.models import Usuario, Vinculo

CPFS_VALIDOS = ["52998224725", "15350946056", "11144477735", "86288366757"]


def _cria(cpf, nome, vinculo, tipo="", lattes=""):
    usuario = Usuario.objects.create_user(
        cpf=cpf,
        email=f"{cpf}@example.com",
        password="senha-Forte-123",
        nome_completo=nome,
        vinculo=vinculo,
        lattes=lattes,
    )
    return Inscricao.objects.create(
        usuario=usuario,
        modalidade=vinculo,
        tipo_servidor=tipo,
        status=StatusInscricao.EM_ANALISE,
        total=10,
    )


class ExportarInscritosLattesTest(TestCase):
    @patch("apps.inscricoes.management.commands.exportar_inscritos_lattes.time.sleep")
    @patch(
        "apps.inscricoes.management.commands.exportar_inscritos_lattes"
        ".buscar_dados_ifgproduz"
    )
    def test_abas_ordenacao_e_subtotais(self, mock_api, _sleep):
        mock_api.return_value = {
            "subtotalA": 10, "subtotalB": 60, "subtotalC": 20.0,
            "subtotalD": 5, "total": 95.0,
        }
        # Ordem proposital embaralhada: apoio técnico "Ana" antes do
        # pesquisador "Zeca" — a aba deve trazer Zeca (pesquisador) primeiro.
        _cria(CPFS_VALIDOS[0], "Ana Apoio", Vinculo.SERVIDOR,
              TipoServidor.APOIO_TECNICO, lattes="http://lattes.cnpq.br/1")
        _cria(CPFS_VALIDOS[1], "Zeca Pesquisador", Vinculo.SERVIDOR,
              TipoServidor.PESQUISADOR)  # sem Lattes
        _cria(CPFS_VALIDOS[2], "Beto Externo", Vinculo.COLABORADOR_EXTERNO,
              TipoServidor.PESQUISADOR, lattes="http://lattes.cnpq.br/2")
        _cria(CPFS_VALIDOS[3], "Carla Estudante", Vinculo.ESTUDANTE)

        saida = Path(tempfile.mkdtemp()) / "teste.xlsx"
        call_command("exportar_inscritos_lattes", saida=str(saida))

        wb = load_workbook(saida)
        self.assertEqual(
            wb.sheetnames, ["Servidores", "Colaboradores Externos", "Estudantes"]
        )

        servidores = list(wb["Servidores"].values)
        # Pesquisador primeiro, mesmo com nome alfabeticamente posterior.
        self.assertEqual(servidores[1][0], "Zeca Pesquisador")
        self.assertEqual(servidores[2][0], "Ana Apoio")
        # Zeca sem Lattes: subtotais vazios, status "sem Lattes".
        self.assertIsNone(servidores[1][6])
        self.assertEqual(servidores[1][-1], "sem Lattes")
        # Ana com Lattes: subtotais preenchidos pela API (mock).
        self.assertEqual(servidores[2][6], 10.0)
        self.assertEqual(servidores[2][10], 95.0)
        self.assertEqual(servidores[2][-1], "ok")

        self.assertEqual(list(wb["Colaboradores Externos"].values)[1][0], "Beto Externo")
        self.assertEqual(list(wb["Estudantes"].values)[1][0], "Carla Estudante")

    def test_banco_vazio_nao_quebra(self):
        saida = Path(tempfile.mkdtemp()) / "vazio.xlsx"
        call_command("exportar_inscritos_lattes", saida=str(saida))
        self.assertFalse(saida.exists())
